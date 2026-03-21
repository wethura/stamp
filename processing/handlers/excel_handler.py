"""Excel 文档处理器"""
from PIL import Image, ImageDraw, ImageFont
from typing import Set, Tuple, Optional
import io

from processing.base import DocumentHandler

# 尝试导入 openpyxl，如果失败则提供友好错误信息
try:
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
    from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def pixels_to_emu(pixels: int) -> int:
    """将像素转换为 EMU (English Metric Units)

    1 英寸 = 914400 EMU
    假设屏幕 DPI 为 96，则 1 像素 = 914400 / 96 = 9525 EMU
    """
    return int(pixels * 9525)


# 预览配置
PREVIEW_ROWS = 25  # 预览显示的行数
PREVIEW_COLS = 10  # 预览显示的列数
CELL_WIDTH = 80    # 预览单元格宽度（像素）
CELL_HEIGHT = 24   # 预览单元格高度（像素）
HEADER_HEIGHT = 40  # 表头高度
TITLE_HEIGHT = 50   # Sheet 标题高度


class ExcelHandler(DocumentHandler):
    """Excel 文档处理器

    支持 .xlsx 格式的读取、预览和盖章导出。
    预览使用 PIL 绘制简化版表格。
    印章使用 AbsoluteAnchor 实现精确位置。
    """

    def __init__(self):
        self._workbook = None
        self._path: Optional[str] = None
        self._sheet_bounds: list = []  # 缓存每个 sheet 的边界

    @classmethod
    def extensions(cls) -> Set[str]:
        return {'.xlsx', '.XLSX'}

    @classmethod
    def can_handle(cls, path: str) -> bool:
        return path.lower().endswith('.xlsx')

    @classmethod
    def display_name(cls) -> str:
        return "Excel 文档"

    def load(self, path: str) -> None:
        if not HAS_OPENPYXL:
            raise RuntimeError("需要安装 openpyxl 库来处理 Excel 文件")

        try:
            self._workbook = load_workbook(path, data_only=True)
            self._path = path
            # 计算每个 sheet 的边界
            self._sheet_bounds = []
            for sheet in self._workbook.worksheets:
                max_row = sheet.max_row
                max_col = sheet.max_column
                self._sheet_bounds.append((max_row, max_col))
        except Exception as e:
            raise RuntimeError(f"无法加载 Excel 文件: {e}")

    def page_count(self) -> int:
        if self._workbook is None:
            return 0
        return len(self._workbook.sheetnames)

    def get_sheet_name(self, index: int) -> str:
        """获取指定索引的 Sheet 名称"""
        if self._workbook is None:
            raise RuntimeError("文档未加载")
        return self._workbook.sheetnames[index]

    def render_page(self, index: int, preview_width: int = 1600) -> Image.Image:
        """渲染 Sheet 为简化预览图像

        使用 PIL 绘制：
        - Sheet 名称作为标题
        - 表格网格线
        - 单元格文本（截断）
        """
        if self._workbook is None:
            raise RuntimeError("文档未加载")

        sheet = self._workbook.worksheets[index]
        sheet_name = self._workbook.sheetnames[index]

        # 获取实际的行列数
        max_row, max_col = self._sheet_bounds[index]
        display_rows = min(max_row, PREVIEW_ROWS)
        display_cols = min(max_col, PREVIEW_COLS)

        # 计算图像尺寸
        img_width = display_cols * CELL_WIDTH + 1
        img_height = TITLE_HEIGHT + HEADER_HEIGHT + display_rows * CELL_HEIGHT + 1

        # 创建图像
        img = Image.new('RGB', (img_width, img_height), color='white')
        draw = ImageDraw.Draw(img)

        # 尝试使用系统字体
        try:
            title_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 18)
            header_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 12)
            cell_font = ImageFont.truetype("/System/Library/Fonts/PingFang.ttc", 11)
        except:
            try:
                title_font = ImageFont.truetype("Arial.ttf", 18)
                header_font = ImageFont.truetype("Arial.ttf", 12)
                cell_font = ImageFont.truetype("Arial.ttf", 11)
            except:
                title_font = ImageFont.load_default()
                header_font = ImageFont.load_default()
                cell_font = ImageFont.load_default()

        # 绘制标题
        draw.text((10, 12), f"Sheet: {sheet_name}", fill='#1a1a1a', font=title_font)

        # 绘制列标头 (A, B, C, ...)
        y_start = TITLE_HEIGHT
        for col in range(display_cols):
            x = col * CELL_WIDTH
            col_letter = self._col_to_letter(col + 1)
            draw.rectangle([x, y_start, x + CELL_WIDTH, y_start + HEADER_HEIGHT],
                          outline='#d0d0d0', fill='#f5f5f5')
            draw.text((x + 4, y_start + 12), col_letter, fill='#666666', font=header_font)

        # 绘制行标头和单元格
        for row in range(display_rows):
            y = y_start + HEADER_HEIGHT + row * CELL_HEIGHT

            # 行标头
            draw.rectangle([0, y, 0, y + CELL_HEIGHT], outline='#d0d0d0', fill='#f5f5f5')
            draw.text((4, y + 5), str(row + 1), fill='#666666', font=header_font)

            # 单元格
            for col in range(display_cols):
                x = col * CELL_WIDTH
                cell = sheet.cell(row=row + 1, column=col + 1)
                value = self._get_cell_text(cell)

                draw.rectangle([x, y, x + CELL_WIDTH, y + CELL_HEIGHT],
                              outline='#e0e0e0', fill='white')
                if value:
                    # 截断文本
                    display_text = self._truncate_text(value, CELL_WIDTH - 6, cell_font)
                    draw.text((x + 3, y + 4), display_text, fill='#333333', font=cell_font)

        # 如果数据被截断，显示提示
        if max_row > PREVIEW_ROWS or max_col > PREVIEW_COLS:
            hint = f"(显示前 {display_rows} 行 × {display_cols} 列，共 {max_row} 行 × {max_col} 列)"
            draw.text((10, img_height - 25), hint, fill='#999999', font=header_font)

        return img

    def get_page_size(self, index: int) -> Tuple[float, float]:
        """获取 Sheet 的虚拟尺寸

        由于 Excel 没有固定的页面尺寸概念，这里返回一个基于内容的虚拟尺寸。
        用于计算印章位置比例。
        """
        if self._workbook is None:
            raise RuntimeError("文档未加载")

        max_row, max_col = self._sheet_bounds[index]
        # 使用像素尺寸作为虚拟页面尺寸
        width = max(PREVIEW_COLS, max_col) * CELL_WIDTH
        height = TITLE_HEIGHT + HEADER_HEIGHT + max(PREVIEW_ROWS, max_row) * CELL_HEIGHT
        return (float(width), float(height))

    def export_with_stamp(
        self,
        output_path: str,
        stamp_img: Image.Image,
        position_ratio: Tuple[float, float],
        stamp_size_ratio: float,
        selected_pages: Set[int]
    ) -> None:
        """导出带印章的 Excel 文件

        使用 AbsoluteAnchor 实现精确的印章位置。
        """
        if self._workbook is None:
            raise RuntimeError("文档未加载")

        x_ratio, y_ratio = position_ratio

        for page_idx in selected_pages:
            if page_idx < 0 or page_idx >= len(self._workbook.worksheets):
                continue

            sheet = self._workbook.worksheets[page_idx]

            # 计算 Sheet 的虚拟尺寸
            max_row, max_col = self._sheet_bounds[page_idx]
            # 使用更大的基础尺寸来匹配 Excel 的实际显示
            base_width = max(1000, max_col * 80)  # 假设每列约 80 像素
            base_height = max(800, max_row * 20)   # 假设每行约 20 像素

            # 计算印章尺寸
            stamp_width = int(base_width * stamp_size_ratio)
            stamp_height = int(stamp_width * stamp_img.height / stamp_img.width)

            # 缩放印章图像
            resample = getattr(Image, "LANCZOS", None) or getattr(Image, "ANTIALIAS")
            scaled_stamp = stamp_img.resize((stamp_width, stamp_height), resample)

            # 将印章保存为字节流
            stamp_bytes = io.BytesIO()
            scaled_stamp.save(stamp_bytes, format='PNG')
            stamp_bytes.seek(0)

            # 创建 Excel 图片对象
            xl_img = XLImage(stamp_bytes)

            # 计算位置（EMU 单位）
            pos_x = pixels_to_emu(int(x_ratio * base_width))
            pos_y = pixels_to_emu(int(y_ratio * base_height))
            ext_x = pixels_to_emu(stamp_width)
            ext_y = pixels_to_emu(stamp_height)

            # 使用 AbsoluteAnchor 设置精确位置
            position = XDRPoint2D(pos_x, pos_y)
            size = XDRPositiveSize2D(ext_x, ext_y)
            xl_img.anchor = AbsoluteAnchor(pos=position, ext=size)

            # 添加图片到工作表
            sheet.add_image(xl_img)

        # 保存文件
        self._workbook.save(output_path)

        # 重新加载以清除添加的图片（避免重复添加）
        if self._path:
            self._workbook = load_workbook(self._path, data_only=True)

    def default_output_extension(self) -> str:
        return '.xlsx'

    def close(self) -> None:
        if self._workbook is not None:
            self._workbook.close()
            self._workbook = None
        self._path = None
        self._sheet_bounds = []

    @staticmethod
    def _col_to_letter(col: int) -> str:
        """将列号转换为 Excel 列字母 (1 -> A, 27 -> AA)"""
        result = ""
        while col > 0:
            col, remainder = divmod(col - 1, 26)
            result = chr(65 + remainder) + result
        return result

    @staticmethod
    def _get_cell_text(cell) -> str:
        """获取单元格的显示文本"""
        if cell.value is None:
            return ""
        value = str(cell.value)
        # 处理浮点数显示
        if isinstance(cell.value, float) and cell.value == int(cell.value):
            return str(int(cell.value))
        return value

    @staticmethod
    def _truncate_text(text: str, max_width: int, font) -> str:
        """截断文本以适应指定宽度"""
        if not text:
            return ""

        # 简单估算：中文字符约等于字体大小，英文约为一半
        # 这里使用保守估计
        max_chars = max_width // 6  # 假设平均字符宽度为 6 像素
        if len(text) <= max_chars:
            return text
        return text[:max_chars - 1] + "…"
